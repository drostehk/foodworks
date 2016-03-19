from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
import pandas as pd


def only_num(string):
    if str(string).isdigit():
        if float(string) != float(string):
            return 0
        else:
            return int(string)
    else:
        return 0
def generate_bread_reprot(bakery_chain, date):
    file_dir = "../../../../Data/FoodLinks-Bread/"
    #file_dir = "/Users/DicksonK/Google Drive/FoodWorks/Data/FoodLinks-Bread/"
    file_name = "foodlinks_bread.csv"
    report_dir = "../../../../Report/bread/"
    #report_dir = "/Users/DicksonK/Google Drive/FoodWorks/Report/bread/"
    template_name = "template/bread_template.xlsx"

    df = pd.read_csv(file_dir + file_name)


    df.columns = ['pair_number', 'dayofweek', 'date', 'amount', 'month', 'picking_days', 'beneficiary_id', 'beneficiary_name', 'shop_id', 'bakery_chain', 'shop_name']

    df_sub = df[['pair_number', 'shop_id', 'month', 'beneficiary_name', 'bakery_chain', 'shop_name', 'amount', 'picking_days']].copy()

    df_sub.amount = df_sub.amount.apply(only_num)
    df_sub.picking_days = df_sub.picking_days.apply(only_num)

    df_result = df_sub[(df_sub.bakery_chain == bakery_chain) & (df_sub.month == date)].groupby(['pair_number', 'shop_id', 'month', 'beneficiary_name', 'bakery_chain', 'shop_name']).sum()

    df_result = df_result.reset_index()
    df_result = df_result[['pair_number', 'shop_id', 'shop_name', 'beneficiary_name', 'amount', 'picking_days']]

    book = load_workbook(report_dir + template_name)
    writer = pd.ExcelWriter(report_dir + bakery_chain.lower() + '_' + date.lower().replace("-", "_") + ".xlsx", engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df_name = pd.DataFrame([[bakery_chain]])
    df_mth = pd.DataFrame([[date]])

    df_avg = pd.DataFrame([[int(df_result.amount.sum() / df_result.picking_days.sum())]])

    df_total = pd.DataFrame([['Grand Total', '', '', '', df_result.amount.sum(), df_result.picking_days.sum()]])

    df_name.to_excel(writer, "Monthly Report", startcol=1, startrow=2, header = False, index = False, merge_cells = False)
    df_mth.to_excel(writer, "Monthly Report", startcol=1, startrow=3, header = False, index = False, merge_cells = False)

    df_avg.to_excel(writer, "Monthly Report", startcol=5, startrow=4, header = False, index = False, merge_cells = False)

    df_result.to_excel(writer, "Monthly Report", startcol=0, startrow=6, header = False, index = False, merge_cells = False)

    df_total.to_excel(writer, "Monthly Report", startcol=0, startrow=(6 + len(df_result)), header = False, index = False, merge_cells = False)

    wb = book
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))


    from openpyxl.styles import fills, PatternFill
    table_fill = PatternFill(patternType=fills.FILL_SOLID, start_color='FFFBCCA5', end_color='FFFBCCA5')

    for row_num in range(len(df_result)):
        for col_num in range(len(df_result.columns)):
            if row_num % 2 == 0:
                ws.cell(row= 7 + row_num, column= 1 + col_num).fill = table_fill
            ws.cell(row= 7 + row_num, column= 1 + col_num).border = thin_border

    total_fill = PatternFill(patternType=fills.FILL_SOLID, start_color='FFF9B784', end_color='FFF9B784')
    for col_num in range(len(df_result.columns)):
        ws.cell(row= 7 + len(df_result), column= 1 + col_num).fill = total_fill
        ws.cell(row= 7 + len(df_result), column= 1 + col_num).border = thin_border


    img = Image('../../../../Report/bread/template/logo.png')
    ws.add_image(img, 'E1')







    writer.save()

bakery_list = ['Yamazaki', "Maxim's", 'Arome', 'Tsui Wah', 'St. Honore', '7-Eleven', 'Hoixe', 'Circle K']

for bakery in bakery_list:
    generate_bread_reprot(bakery, 'Dec-15')